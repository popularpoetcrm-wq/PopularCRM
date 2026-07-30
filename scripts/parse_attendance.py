#!/usr/bin/env python3
"""Phase 2: Excel date columns → sessions + attendance JSON."""

from __future__ import annotations

import json
import re
import sys
from datetime import date, datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

# reuse phase-1 helpers
sys.path.insert(0, str(Path(__file__).resolve().parent))
from parse_real_tables import (  # noqa: E402
    ROOT,
    TABLES,
    TENANT,
    clean_name,
    detect_has_size_col,
    name_key,
    parse_schedule,
    parse_workbook,
    uid,
)

try:
    import openpyxl
except ImportError as e:
    raise SystemExit("Need openpyxl (.venv-xlsx)") from e

WARSAW = ZoneInfo("Europe/Warsaw")
OUT = ROOT / "scripts" / "data" / "real-tables-phase2-attendance.json"


def header_dates(header: tuple) -> list[tuple[int, date]]:
    out: list[tuple[int, date]] = []
    for i, v in enumerate(header):
        if isinstance(v, datetime):
            out.append((i, v.date()))
        elif isinstance(v, date):
            out.append((i, v))
    return out


def interpret_mark(raw: object) -> dict | None:
    """Map journal cell → attendance status."""
    if raw is None or raw == "":
        return None
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        # lesson counter number = present
        return {
            "status": "present",
            "attendance_type": "regular",
            "comment": str(int(raw) if float(raw).is_integer() else raw),
        }

    s = str(raw).strip()
    if not s:
        return None
    compact = re.sub(r"\s+", "", s).lower()

    if compact in {"проба", "proba", "trial", "пр"}:
        return {"status": "present", "attendance_type": "regular", "comment": "trial"}

    if "*+" in compact or "+*" in compact:
        return {"status": "present", "attendance_type": "makeup", "comment": s}

    # absent marks: -, $-, -$
    if re.fullmatch(r"\$*-+\$*", compact) or compact in {"н", "abs"}:
        return {
            "status": "absent_notified",
            "attendance_type": "regular",
            "comment": s,
        }

    # anything with a digit (1$, $5, 12.0, 5(420)$ …) = was there / counted
    if re.search(r"\d", compact):
        return {"status": "present", "attendance_type": "regular", "comment": s}

    # leftover symbols — treat as present note
    return {"status": "present", "attendance_type": "regular", "comment": s}


def session_times(day: date, sched: dict) -> tuple[datetime, datetime]:
    start_s = sched.get("start_time") or "18:00"
    dur = int(sched.get("duration_minutes") or 90)
    hh, mm = (int(x) for x in start_s.split(":")[:2])
    start = datetime(day.year, day.month, day.day, hh, mm, tzinfo=WARSAW)
    end = start + timedelta(minutes=dur)
    return start, end


def parse_attendance_workbook(
    path: Path,
    brand: str,
    *,
    is_minor: bool,
    has_size_col_default: bool,
) -> dict:
    # ensure same group/person/enrollment ids as phase 1
    meta = parse_workbook(
        path,
        brand,
        has_size_col=has_size_col_default,
        is_minor=is_minor,
        skip={"абонемент"},
    )
    groups_by_sheet = {g["source_sheet"]: g for g in meta["groups"]}
    enroll_by_key = {
        (e["group_id"], e["student_person_id"]): e for e in meta["enrollments"]
    }

    wb = openpyxl.load_workbook(path, data_only=True)
    sessions: dict[str, dict] = {}
    attendance: list[dict] = []
    skipped_marks: dict[str, int] = {}

    for sheet_name, group in groups_by_sheet.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        dates = header_dates(header)
        if not dates:
            continue

        sched = group.get("schedule") or {}
        sheet_has_size = (
            False if brand == "kids" else detect_has_size_col(ws)
        )

        for row in ws.iter_rows(min_row=2, values_only=True):
            name = clean_name(row[0] if row else None)
            if not name:
                continue
            pid = uid("person", brand, name_key(row[0]) or name.casefold())
            eid_key = (group["id"], pid)
            enr = enroll_by_key.get(eid_key)
            if not enr:
                continue

            for col_i, day in dates:
                if col_i >= len(row):
                    continue
                mark = interpret_mark(row[col_i])
                if not mark:
                    continue

                start, end = session_times(day, sched)
                sid = uid("session", group["id"], day.isoformat())
                if sid not in sessions:
                    now = datetime.now(tz=WARSAW)
                    sessions[sid] = {
                        "id": sid,
                        "tenant_id": TENANT,
                        "group_id": group["id"],
                        "brand_id": brand,
                        "starts_at": start.isoformat(),
                        "ends_at": end.isoformat(),
                        "status": "completed" if end < now else "scheduled",
                        "source_sheet": sheet_name,
                        "session_date": day.isoformat(),
                    }

                aid = uid("att", sid, enr["id"])
                attendance.append(
                    {
                        "id": aid,
                        "tenant_id": TENANT,
                        "session_id": sid,
                        "enrollment_id": enr["id"],
                        "student_person_id": pid,
                        "attendance_type": mark["attendance_type"],
                        "status": mark["status"],
                        "comment": mark.get("comment"),
                    }
                )

        # collect unknown? interpret always returns something for non-empty
        _ = sheet_has_size
        _ = skipped_marks

    return {
        "brand": brand,
        "file": path.name,
        "sessions": list(sessions.values()),
        "attendance": attendance,
        "totals": {
            "sessions": len(sessions),
            "attendance": len(attendance),
            "present": sum(1 for a in attendance if a["status"] == "present"),
            "absent": sum(1 for a in attendance if a["status"] != "present"),
        },
    }


def main() -> None:
    poet = parse_attendance_workbook(
        TABLES / "Популярный поэт.xlsx",
        "poet",
        is_minor=False,
        has_size_col_default=True,
    )
    kids = parse_attendance_workbook(
        TABLES / "Идея.xlsx",
        "kids",
        is_minor=True,
        has_size_col_default=False,
    )

    # dedupe attendance by id
    att_map = {a["id"]: a for a in poet["attendance"] + kids["attendance"]}
    sess_map = {s["id"]: s for s in poet["sessions"] + kids["sessions"]}

    out = {
        "tenant_id": TENANT,
        "phase": 2,
        "note": "Sessions from date headers; marks → present / absent_notified / makeup",
        "poet": poet,
        "kids": kids,
        "sessions": list(sess_map.values()),
        "attendance": list(att_map.values()),
        "totals": {
            "sessions": len(sess_map),
            "attendance": len(att_map),
            "present": sum(1 for a in att_map.values() if a["status"] == "present"),
            "absent": sum(1 for a in att_map.values() if a["status"] != "present"),
            "poet": poet["totals"],
            "kids": kids["totals"],
        },
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(out["totals"], ensure_ascii=False, indent=2))
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main()
