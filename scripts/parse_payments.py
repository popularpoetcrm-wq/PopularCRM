#!/usr/bin/env python3
"""Parse paid 4-lesson package cycles from the real studio workbooks.

Confirmed studio legend:
- Column ``$`` / price = full package price in PLN.
- A ``$`` inside a dated attendance cell starts a new 4-lesson package.
- The number can be continuous (``1$``, ``5$``, ``9$`` ...); it is not the
  package size and it is not the amount paid.
- Every dated ``$`` marker through 2026-07-30 confirms that the full package
  was paid.

The generated JSON is only a review artifact. Writing it to Supabase requires
the separate ``node scripts/import-payments.mjs --apply`` command.
"""

from __future__ import annotations

import json
import re
import uuid
from datetime import date, datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import openpyxl

from parse_attendance import interpret_mark
from parse_real_tables import clean_name, detect_has_size_col, name_key

ROOT = Path(__file__).resolve().parents[1]
TABLES = ROOT / "Таблицы_реальные"
TENANT = "aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa"
NS = uuid.UUID(TENANT)
SKIP = {"абонемент"}
PAID_THROUGH = date(2026, 7, 30)
WARSAW = ZoneInfo("Europe/Warsaw")


def uid(*parts: str) -> str:
    return str(uuid.uuid5(NS, ":".join(parts)))


def as_price(v: object) -> float | None:
    if v is None or v == "" or v == "-":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def cell_has_pay_mark(v: object) -> bool:
    return v is not None and "$" in str(v)


def header_dates(header: tuple) -> list[tuple[int, date]]:
    dates: list[tuple[int, date]] = []
    for col_index, raw in enumerate(header):
        if isinstance(raw, datetime):
            dates.append((col_index, raw.date()))
        elif isinstance(raw, date):
            dates.append((col_index, raw))
    return dates


def direction_key(title: str) -> str:
    blob = title.lower()
    if "идея" in blob or re.search(r"\d\s*групп", blob):
        return "kids"
    if "импро" in blob:
        return "impro"
    if "актёр" in blob or "актер" in blob:
        return "acting"
    if "воскресн" in blob or "школ" in blob:
        return "school"
    if "спектакл" in blob:
        return "show"
    if "play" in blob:
        return "playback"
    return "other"


def paid_at(day: date) -> str:
    """Keep the source date without inventing a bank transaction time."""
    return datetime(day.year, day.month, day.day, 12, tzinfo=WARSAW).isoformat()


def consumes_regular_credit(raw: object) -> bool:
    mark = interpret_mark(raw)
    return bool(
        mark
        and mark["status"] == "present"
        and mark["attendance_type"] == "regular"
    )


def enrollment_statuses() -> dict[str, str]:
    path = ROOT / "scripts" / "data" / "real-tables-phase1.json"
    if not path.exists():
        return {}
    phase1 = json.loads(path.read_text(encoding="utf-8"))
    rows = [
        *phase1.get("poet", {}).get("enrollments", []),
        *phase1.get("kids", {}).get("enrollments", []),
    ]
    return {row["id"]: row.get("status", "ended") for row in rows}


def parse_sheet(
    ws,
    *,
    brand: str,
    sheet_name: str,
    path_name: str,
    statuses: dict[str, str],
) -> tuple[list[dict], list[dict], list[dict]]:
    has_size = False if brand == "kids" else detect_has_size_col(ws)
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    dates = header_dates(header)
    payments: list[dict] = []
    packages: list[dict] = []
    skipped: list[dict] = []

    for row_index, row in enumerate(
        ws.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        raw_name = row[0] if row else None
        key = name_key(raw_name)
        if not key:
            continue
        display = clean_name(raw_name) or key
        if display.casefold() in {"майки", "майка", "футболки", "импро", "имя", "name"}:
            continue

        if has_size:
            price = as_price(row[2] if len(row) > 2 else None)
            b_num = as_price(row[1] if len(row) > 1 else None)
            if price is None and b_num is not None and b_num <= 100:
                continue
        else:
            price = as_price(row[1] if len(row) > 1 else None)

        pid = uid("person", brand, key)
        gid = uid("group", brand, sheet_name)
        eid = uid("enroll", brand, sheet_name, key)
        cycle_cells: list[tuple[int, date, object]] = [
            (col_index, day, row[col_index])
            for col_index, day in dates
            if col_index < len(row) and cell_has_pay_mark(row[col_index])
        ]
        if not cycle_cells:
            continue

        if price is None:
            skipped.append(
                {
                    "brand_id": brand,
                    "source_sheet": sheet_name,
                    "row": row_index,
                    "full_name": display,
                    "reason": "payment marker exists but package price is missing",
                    "markers": [
                        {"date": day.isoformat(), "value": str(value)}
                        for _, day, value in cycle_cells
                    ],
                }
            )
            continue

        amount = round(price, 2)
        for cycle_index, (col_index, day, value) in enumerate(cycle_cells, start=1):
            is_confirmed_paid = day <= PAID_THROUGH
            if not is_confirmed_paid:
                skipped.append(
                    {
                        "brand_id": brand,
                        "source_sheet": sheet_name,
                        "row": row_index,
                        "full_name": display,
                        "reason": "payment marker is after the confirmed cutoff date",
                        "markers": [
                            {"date": day.isoformat(), "value": str(value)}
                        ],
                    }
                )
                continue
            payments.append(
                {
                    "id": uid(
                        "payment",
                        brand,
                        sheet_name,
                        key,
                        day.isoformat(),
                        str(col_index),
                    ),
                    "tenant_id": TENANT,
                    "provider": "other",
                    "provider_session_id": (
                        f"import:{brand}:{sheet_name}:{key}:"
                        f"{day.isoformat()}:{col_index}"
                    ),
                    "payer_person_id": pid,
                    "enrollment_id": eid,
                    "student_person_id": pid,
                    "group_id": gid,
                    "amount": amount,
                    "amount_paid": amount,
                    "due_at": day.isoformat(),
                    "paid_at": paid_at(day),
                    "currency": "PLN",
                    "status": "paid",
                    "payment_method": "cash",
                    "description": (
                        f"Абонемент 4 занятия · {display} · "
                        f"цикл от {day.strftime('%d.%m.%Y')} · "
                        f"{path_name}/{sheet_name}"
                    ),
                    "price_hint": amount,
                    "cycle_date": day.isoformat(),
                    "cycle_index": cycle_index,
                    "source_cell_value": str(value),
                    "full_name": display,
                    "brand_id": brand,
                    "source_sheet": sheet_name,
                    "direction": direction_key(sheet_name),
                    "requires_review": False,
                }
            )

        # The package balance is useful only for a person who still attends.
        # Start from the last paid cycle at/before the cutoff and count regular
        # visits after it. Absence ($-) and makeup ($*+) confirm the package
        # payment but do not consume a regular package credit.
        if statuses.get(eid) != "active":
            continue
        confirmed_cycles = [cell for cell in cycle_cells if cell[1] <= PAID_THROUGH]
        if not confirmed_cycles:
            continue
        last_col, last_day, last_value = confirmed_cycles[-1]
        visits = [
            {
                "date": day.isoformat(),
                "value": str(row[col_index]),
            }
            for col_index, day in dates
            if last_col <= col_index
            and day <= PAID_THROUGH
            and col_index < len(row)
            and consumes_regular_credit(row[col_index])
        ]
        raw_consumed = len(visits)
        consumed = min(4, raw_consumed)
        packages.append(
            {
                "enrollment_id": eid,
                "student_person_id": pid,
                "group_id": gid,
                "brand_id": brand,
                "full_name": display,
                "source_sheet": sheet_name,
                "cycle_started_at": last_day.isoformat(),
                "source_cell_value": str(last_value),
                "credits_total": 4,
                "credits_consumed": consumed,
                "credits_available": max(0, 4 - consumed),
                "regular_visits_since_cycle_start": visits,
                "needs_review": raw_consumed > 4,
                "review_reason": (
                    "more than four regular visits after the latest payment marker"
                    if raw_consumed > 4
                    else None
                ),
            }
        )

    return payments, packages, skipped


def parse_file(
    path: Path,
    brand: str,
    statuses: dict[str, str],
) -> tuple[list[dict], list[dict], list[dict]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    payments: list[dict] = []
    packages: list[dict] = []
    skipped: list[dict] = []
    for sheet_name in wb.sheetnames:
        if sheet_name.lower() in SKIP:
            continue
        sheet_payments, sheet_packages, sheet_skipped = parse_sheet(
            wb[sheet_name],
            brand=brand,
            sheet_name=sheet_name,
            path_name=path.name,
            statuses=statuses,
        )
        payments.extend(sheet_payments)
        packages.extend(sheet_packages)
        skipped.extend(sheet_skipped)
    wb.close()
    return payments, packages, skipped


def main() -> None:
    statuses = enrollment_statuses()
    payments: list[dict] = []
    packages: list[dict] = []
    skipped: list[dict] = []

    for filename, brand in [
        ("Популярный поэт.xlsx", "poet"),
        ("Идея.xlsx", "kids"),
    ]:
        path = TABLES / filename
        if not path.exists():
            continue
        file_payments, file_packages, file_skipped = parse_file(path, brand, statuses)
        payments.extend(file_payments)
        packages.extend(file_packages)
        skipped.extend(file_skipped)

    payments = list({row["id"]: row for row in payments}.values())
    packages = list(
        {row["enrollment_id"]: row for row in packages}.values()
    )
    paid = [row for row in payments if row["status"] == "paid"]
    package_reviews = [row for row in packages if row["needs_review"]]
    missing_price = [
        row
        for row in skipped
        if row["reason"] == "payment marker exists but package price is missing"
    ]
    future_markers = [
        row
        for row in skipped
        if row["reason"] == "payment marker is after the confirmed cutoff date"
    ]

    totals = {
        "payment_events": len(payments),
        "unique_people": len({row["student_person_id"] for row in payments}),
        "paid_rows_through_2026_07_30": len(paid),
        "revenue_paid": round(sum(row["amount_paid"] for row in paid), 2),
        "active_packages": len(packages),
        "active_package_credits_available": sum(
            row["credits_available"] for row in packages
        ),
        "package_balances_needing_review": len(package_reviews),
        "markers_skipped_missing_price": sum(
            len(row["markers"]) for row in missing_price
        ),
        "future_markers_needing_review": sum(
            len(row["markers"]) for row in future_markers
        ),
    }

    out = {
        "phase": 3,
        "note": (
            "Each dated $ marker starts and confirms payment of a full "
            "4-lesson package; markers through 2026-07-30 are paid"
        ),
        "cutoff_date": PAID_THROUGH.isoformat(),
        "tenant_id": TENANT,
        "totals": totals,
        "payments": payments,
        "current_packages_preview": packages,
        "skipped_for_review": skipped,
    }
    dest = ROOT / "scripts" / "data" / "real-tables-phase3-payments.json"
    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"wrote": str(dest), **totals}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
