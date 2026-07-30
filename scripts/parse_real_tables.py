#!/usr/bin/env python3
"""Parse Таблицы_реальные → JSON for Supabase import (phase 1: groups + students)."""

from __future__ import annotations

import json
import re
import uuid
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError as e:
    raise SystemExit("Need openpyxl (use .venv-xlsx)") from e

ROOT = Path(__file__).resolve().parents[1]
TABLES = ROOT / "Таблицы_реальные"
TENANT = "aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa"
NS = uuid.UUID(TENANT)
PLAN_ID = "bbbbbbbb-bbbb-bbbb-bbbb-bbbbbbbbbbbb"

SKIP_SHEETS = {
    "абонемент",  # payments — later phase
}

# Sheets in poet workbook that are kids-ish / special — still brand poet unless noted
WEEKDAY = {
    "пн": 1,
    "вт": 2,
    "ср": 3,
    "чт": 4,
    "пт": 5,
    "сб": 6,
    "вс": 0,
}

WEEKDAY_RU = {
    0: "Воскресенье",
    1: "Понедельник",
    2: "Вторник",
    3: "Среда",
    4: "Четверг",
    5: "Пятница",
    6: "Суббота",
}

DIRECTION_CODE = {
    "impro": "impro",
    "импро": "impro",
    "импроверты": "impro",
    "acting": "acting",
    "актёрка": "acting",
    "актерка": "acting",
    "актёр": "acting",
    "актер": "acting",
    "show": "show",
    "спектакль": "show",
    "спектакл": "show",
    "playback": "playback",
    "play-back": "playback",
    "sunday_school": "school",
    "воскресная школа": "school",
}

DIRECTION_RU = {
    "impro": "Импровизация",
    "acting": "Актёрское мастерство",
    "show": "Спектакль",
    "playback": "Play-back",
    "school": "Воскресная школа",
    "kids": "Детская студия",
}


def uid(*parts: str) -> str:
    return str(uuid.uuid5(NS, ":".join(parts)))


def name_key(raw: object) -> str | None:
    """Stable identity key — keeps emoji so «Маша👩🏽» ≠ «Маша»."""
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    s = re.sub(r"\$+$", "", s).strip()
    s = re.sub(r"\s+", " ", s)
    return s.casefold() if s else None


def clean_name(raw: object) -> str | None:
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    # trailing $ = paid marker on name in some rows
    s = re.sub(r"\$+$", "", s).strip()
    # strip emoji / variation selectors / dingbats (display only)
    s = re.sub(
        r"[\U0001F200-\U0001FAFF\U00002700-\U000027BF\U0001F1E0-\U0001F1FF"
        r"\U0000FE0F\U0000200D\U00002600-\U000026FF]+",
        "",
        s,
    ).strip()
    s = re.sub(r"\s+", " ", s)
    # leftover symbols like ^-^
    s = re.sub(r"[\^~]+", "", s).strip()
    if not s:
        return None
    return s


def unique_display_name(base: str, existing: set[str]) -> str:
    if base not in existing:
        return base
    n = 2
    while f"{base} ({n})" in existing:
        n += 1
    return f"{base} ({n})"


def row_has_journal_marks(row: tuple, date_start_col: int = 4) -> bool:
    """True if any date-column cell has a mark (not just empty)."""
    for cell in row[date_start_col:]:
        if cell is None or cell == "":
            continue
        if isinstance(cell, (datetime, date)):
            continue
        return True
    return False


def as_date(v: object) -> str | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return None


def as_price(v: object) -> float | None:
    if v is None or v == "" or v == "-":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", ".")
    if s in {"-", "—", ""}:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def as_size(v: object) -> str | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return None  # price leaked into size column
    s = str(v).strip().upper().replace(" ", "")
    if not s or s == "-":
        return None
    if re.fullmatch(r"[\d.,]+", s):
        return None
    # normalize latin/cyrillic M
    s = s.replace("М", "M").replace("Х", "X")
    if not re.fullmatch(r"X{0,3}S|X{0,2}M|X{0,3}L|[0-9]?X{0,2}[SML]", s):
        # allow common sizes only
        if s not in {"XS", "S", "M", "L", "XL", "XXL", "XXXL", "3XL", "4XL"}:
            return None
    return s[:8]


def detect_has_size_col(ws) -> bool:
    """Poet sheets: B1='R' means size. B1='$' means price (no size), like Воскресная школа."""
    b1 = ws.cell(1, 2).value
    if b1 is None:
        return False
    marker = str(b1).strip().upper()
    if marker in {"$", "ЦЕНА", "PRICE", "PLN"}:
        return False
    if marker in {"R", "SIZE", "РАЗМЕР", "ФУТБОЛКА"}:
        return True
    # peek first data cells: size-like vs price-like
    for row in ws.iter_rows(min_row=2, max_row=8, min_col=2, max_col=2, values_only=True):
        cell = row[0]
        if cell is None or cell == "":
            continue
        if as_size(cell):
            return True
        if as_price(cell) is not None:
            return False
    return False


def normalize_direction(raw: str | None, sheet_name: str = "") -> str | None:
    blob = f"{raw or ''} {sheet_name}".lower()
    for key, code in DIRECTION_CODE.items():
        if key in blob:
            return code
    if raw and raw.strip() and not re.fullmatch(r"[\d:.\-\s–]+", raw.strip()):
        # drop weekday-only headers like "Понедельник", "Пятница", "Суббота"
        if raw.strip().lower() in {v.lower() for v in WEEKDAY_RU.values()} | {
            "понедельник",
            "вторник",
            "среда",
            "четверг",
            "пятница",
            "суббота",
            "воскресенье",
            "четверг импро",
        }:
            return None
        return "other"
    return None


def parse_time_range(blob: str) -> tuple[str | None, str | None, int]:
    """Return start HH:MM, end HH:MM, duration minutes."""
    compact = blob.replace(" ", "")
    # 18:00-20:00 / 11:30–13:30
    m = re.search(
        r"(\d{1,2})[:.](\d{2})\s*[-–]\s*(\d{1,2})[:.](\d{2})",
        blob,
    )
    if m:
        h1, m1, h2, m2 = (int(x) for x in m.groups())
        start = f"{h1:02d}:{m1:02d}"
        end = f"{h2:02d}:{m2:02d}"
        dur = max(30, (h2 * 60 + m2) - (h1 * 60 + m1))
        return start, end, dur

    # 1800-2000 / 1830-2030
    m2 = re.search(r"(\d{3,4})\s*[-–]\s*(\d{3,4})", compact)
    if m2:
        a, b = m2.group(1), m2.group(2)
        if len(a) == 3:
            a = "0" + a
        if len(b) == 3:
            b = "0" + b
        start = f"{a[:2]}:{a[2:]}"
        end = f"{b[:2]}:{b[2:]}"
        dur = max(30, (int(b[:2]) * 60 + int(b[2:])) - (int(a[:2]) * 60 + int(a[2:])))
        return start, end, dur

    # single time 1900 or 19:00 → +2h default
    m3 = re.search(r"(?:^|[^\d])(\d{3,4})(?:[^\d]|$)", compact)
    if m3:
        a = m3.group(1)
        if len(a) == 3:
            a = "0" + a
        if len(a) == 4:
            h, mi = int(a[:2]), int(a[2:])
            if 0 <= h <= 23 and 0 <= mi <= 59:
                start = f"{h:02d}:{mi:02d}"
                end_m = h * 60 + mi + 120
                end = f"{end_m // 60:02d}:{end_m % 60:02d}"
                return start, end, 120
    m4 = re.search(r"(\d{1,2})[:.](\d{2})", blob)
    if m4:
        h, mi = int(m4.group(1)), int(m4.group(2))
        start = f"{h:02d}:{mi:02d}"
        end_m = h * 60 + mi + 120
        end = f"{end_m // 60:02d}:{end_m % 60:02d}"
        return start, end, 120

    return None, None, 90


def parse_schedule(sheet_name: str, header_a1: object) -> dict:
    name_l = sheet_name.lower()
    header = str(header_a1).strip() if header_a1 is not None else ""
    header_l = header.lower()

    weekday = None
    for key, num in WEEKDAY.items():
        if name_l.startswith(key) or re.search(rf"(^|[\s\-]){key}([\s\-]|$)", name_l):
            weekday = num
            break
    for word, num in (
        ("понедельник", 1),
        ("вторник", 2),
        ("сред", 3),
        ("четверг", 4),
        ("пятниц", 5),
        ("суббот", 6),
        ("воскрес", 0),
    ):
        if word in name_l or word in header_l:
            weekday = num
            break

    blob = f"{sheet_name} {header}"
    start_time, end_time, duration = parse_time_range(blob)

    direction_raw = header if header and not re.fullmatch(r"[\d:.\-\s–]+", header) else None
    # kids headers often "Четверг 16:00-17:30" / "Макс 15:00" — not a direction label
    if direction_raw and re.search(
        r"(понедельник|вторник|сред|четверг|пятниц|суббот|воскрес|\d{1,2}[:.]\d{2}|\d{3,4})",
        direction_raw.lower(),
    ):
        # keep only if clearly a course name without schedule noise... else None
        if not any(k in direction_raw.lower() for k in ("импро", "актёр", "актер", "спектакл", "play")):
            direction_raw = None
    direction = normalize_direction(direction_raw, sheet_name)

    return {
        "weekday": weekday,
        "start_time": start_time,
        "end_time": end_time,
        "duration_minutes": duration,
        "direction": direction,
    }


def format_time_range(sched: dict) -> str | None:
    start = sched.get("start_time")
    end = sched.get("end_time")
    if start and end:
        return f"{start}–{end}"
    if start:
        return start
    return None


def group_title(brand: str, sheet: str, header_a1: object, sched: dict) -> str:
    day = WEEKDAY_RU.get(sched["weekday"]) if sched.get("weekday") is not None else None
    time_part = format_time_range(sched)
    direction = sched.get("direction")

    if brand == "kids":
        # "1 группа" → группа 1
        m = re.search(r"(\d+)\s*групп", sheet.lower())
        group_no = m.group(1) if m else sheet.strip()
        bits = ["Идея"]
        if day:
            bits.append(day)
        if time_part:
            bits.append(time_part)
        title = " · ".join(bits)
        if m:
            title = f"{title} — группа {group_no}"
        elif direction and direction not in title:
            title = f"{title} — {direction}"
        # teacher hint from header like "Макс 15:00-16:30"
        header = str(header_a1).strip() if header_a1 else ""
        teacher = re.match(r"^([A-Za-zА-Яа-яЁё]+)\s+\d", header)
        if teacher and teacher.group(1).lower() not in {
            "четверг",
            "суббота",
            "понедельник",
            "вторник",
            "среда",
            "пятница",
            "воскресенье",
        }:
            title = f"{title} ({teacher.group(1)})"
        return title

    # poet: «Вторник 18:00–20:00 — Импровизация»
    head = " ".join(x for x in (day, time_part) if x)
    header = str(header_a1).strip() if header_a1 else ""
    # keep studio nickname when it adds meaning (Импроверты ≠ generic Импро)
    nickname = None
    if header and header.lower() not in {
        "импро",
        "импровизация",
        "актёрка",
        "актерка",
        "актёрское мастерство",
    }:
        if not re.fullmatch(r"[\d:.\-\s–]+", header) and header.lower() not in {
            v.lower() for v in WEEKDAY_RU.values()
        }:
            if "импроверт" in header.lower():
                nickname = "Импроверты"
            elif header.lower() in {"спектакль"} or direction == header:
                nickname = None
            elif direction and header.lower() not in direction.lower() and len(header) < 40:
                # e.g. "20:00-22:00" already stripped; keep short labels
                if not re.search(r"\d", header):
                    nickname = header

    label = DIRECTION_RU.get(direction or "", direction)
    if nickname and nickname != label and nickname.lower() not in {
        "импро",
        "импровизация",
        "актёрка",
        "актерка",
        "четверг импро",
    }:
        if "импроверт" in nickname.lower():
            label = f"{label} · Импроверты" if label else "Импроверты"
        elif label and len(nickname) < 40 and not re.search(r"\d", nickname):
            label = f"{label} · {nickname}"
        elif not label:
            label = nickname

    if label and head:
        return f"{head} — {label}"
    if label:
        return label
    if head:
        return head
    return sheet.strip()


def parse_workbook(
    path: Path,
    brand: str,
    *,
    has_size_col: bool,
    is_minor: bool,
    skip: set[str] | None = None,
) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    skip = {s.lower() for s in (skip or set())} | SKIP_SHEETS
    groups: list[dict] = []
    persons: dict[str, dict] = {}
    enrollments: list[dict] = []

    for sheet_name in wb.sheetnames:
        if sheet_name.lower() in skip:
            continue
        ws = wb[sheet_name]
        header_a1 = ws.cell(1, 1).value
        sched = parse_schedule(sheet_name, header_a1)
        gid = uid("group", brand, sheet_name)
        title = group_title(brand, sheet_name, header_a1, sched)
        groups.append(
            {
                "id": gid,
                "brand_id": brand,
                "title": title,
                "direction": sched.get("direction"),
                "source_sheet": sheet_name,
                "schedule": sched,
                "capacity": 16 if brand == "poet" else 12,
            }
        )

        # auto: size column if header B is R; Воскресная школа has B=$
        sheet_has_size = has_size_col if brand == "kids" else detect_has_size_col(ws)
        if brand == "kids":
            sheet_has_size = False

        # poet: A=name B=size C=price D=birthday  OR  A=name B=price C=birthday
        # Blank row(s) after the first block → people below are inactive («не ходят»).
        enrollment_status = "active"
        saw_person = False
        for row in ws.iter_rows(min_row=2, values_only=True):
            raw0 = row[0] if row else None
            name = clean_name(raw0)
            if not name:
                if saw_person:
                    enrollment_status = "ended"
                continue

            # Section headers / noise
            if name.casefold() in {
                "майки",
                "майка",
                "футболки",
                "размер",
                "имена",
                "ученики",
            }:
                enrollment_status = "ended"
                continue

            if sheet_has_size:
                size = as_size(row[1] if len(row) > 1 else None)
                price = as_price(row[2] if len(row) > 2 else None)
                bday = as_date(row[3] if len(row) > 3 else None)
                # Inventory block under «Майки»: name + number in B, no price/birthday
                b_num = as_price(row[1] if len(row) > 1 else None)
                if (
                    size is None
                    and price is None
                    and bday is None
                    and b_num is not None
                    and b_num <= 100
                ):
                    continue
                date_col = 4
            else:
                size = None
                price = as_price(row[1] if len(row) > 1 else None)
                bday = as_date(row[2] if len(row) > 2 else None)
                date_col = 3

            # After the gap: bare name duplicates without price/marks = notes, not people
            if (
                enrollment_status == "ended"
                and size is None
                and price is None
                and bday is None
                and not row_has_journal_marks(row, date_col)
            ):
                continue

            saw_person = True

            key = name_key(raw0)
            if not key:
                continue
            display = clean_name(raw0) or key
            # stable person key within brand (emoji kept in key)
            pid = uid("person", brand, key)
            if pid not in persons:
                existing_names = {p["full_name"] for p in persons.values()}
                persons[pid] = {
                    "id": pid,
                    "full_name": unique_display_name(display, existing_names),
                    "birth_date": bday,
                    "tshirt_size": size,
                    "is_minor": is_minor,
                    "brand_hint": brand,
                    "source": path.name,
                }
            else:
                # fill missing attrs; for active block prefer this sheet's size/bday
                p = persons[pid]
                if p.get("tshirt_size") and not as_size(p["tshirt_size"]):
                    p["tshirt_size"] = None
                if bday and (not p.get("birth_date") or enrollment_status == "active"):
                    p["birth_date"] = bday
                if size and (not p.get("tshirt_size") or enrollment_status == "active"):
                    p["tshirt_size"] = size

            eid = uid("enroll", brand, sheet_name, key)
            # one enrollment per person/group (sheet may list same name twice)
            existing = next((x for x in enrollments if x["id"] == eid), None)
            if not existing:
                enrollments.append(
                    {
                        "id": eid,
                        "student_person_id": pid,
                        "group_id": gid,
                        "brand_id": brand,
                        "status": enrollment_status,
                        "price_hint": price,
                        "tags": [f"price:{int(price)}"] if price else [],
                    }
                )
            else:
                # Prefer active if listed in the top block later/earlier
                if existing.get("status") != "active" and enrollment_status == "active":
                    existing["status"] = "active"
                if price and not existing.get("price_hint"):
                    existing["price_hint"] = price
                    existing["tags"] = [f"price:{int(price)}"]

    apply_sheet_fixes(groups, brand)

    return {
        "brand": brand,
        "file": path.name,
        "groups": groups,
        "persons": list(persons.values()),
        "enrollments": enrollments,
    }


def apply_sheet_fixes(groups: list[dict], brand: str) -> None:
    if brand != "poet":
        for g in groups:
            g["direction"] = g.get("direction") or "kids"
            g["schedule"]["direction"] = g["direction"]
        return

    sheet_fixes = {
        "вс 1230-1430": {
            "direction": "impro",
            "title": "Воскресенье 12:30–14:30 — Импровизация",
        },
        "play-back": {"direction": "playback", "title": "Play-back"},
        "чт 2000-2200": {
            "direction": "acting",
            "title": "Четверг 20:00–22:00 — Актёрское мастерство",
        },
        "пн 1830-2030": {
            "direction": "impro",
            "title": "Понедельник 18:30–20:30 — Импровизация",
        },
        "сб 1700-1900": {
            "direction": "acting",
            "title": "Суббота 17:00–19:00 — Актёрское мастерство",
        },
        "пт 1900-2100": {
            "direction": "acting",
            "title": "Пятница 19:00–21:00 — Актёрское мастерство",
        },
        "вс 1800-2000": {
            "direction": "show",
            "title": "Воскресенье 18:00–20:00 — Спектакль",
        },
        "воскресная школа": {
            "direction": "school",
            "title": "Воскресенье — Воскресная школа",
        },
        "вт 1800-2000": {
            "direction": "impro",
            "title": "Вторник 18:00–20:00 — Импровизация · Импроверты",
        },
    }
    for g in groups:
        sheet = g["source_sheet"].lower()
        fix = sheet_fixes.get(sheet)
        if fix:
            g["direction"] = fix["direction"]
            g["schedule"]["direction"] = fix["direction"]
            if fix.get("title"):
                g["title"] = fix["title"]
            if "воскресная" in sheet:
                g["schedule"]["weekday"] = 0
        elif "воскресная" in sheet:
            g["direction"] = "school"
            g["schedule"]["direction"] = "school"
            g["schedule"]["weekday"] = 0
            g["title"] = "Воскресенье — Воскресная школа"
        elif not g.get("direction"):
            # fallback from already-normalized schedule
            g["direction"] = (g.get("schedule") or {}).get("direction")


def main() -> None:
    poet_path = TABLES / "Популярный поэт.xlsx"
    idea_path = TABLES / "Идея.xlsx"
    if not poet_path.exists() or not idea_path.exists():
        raise SystemExit(f"Missing xlsx in {TABLES}")

    poet = parse_workbook(
        poet_path,
        "poet",
        has_size_col=True,
        is_minor=False,
        skip={"абонемент"},
    )
    kids = parse_workbook(
        idea_path,
        "kids",
        has_size_col=False,
        is_minor=True,
    )

    out = {
        "tenant_id": TENANT,
        "plan_id": PLAN_ID,
        "phase": 1,
        "note": "Phase 1: groups + persons + enrollments. Blank-row gap → ended. Attendance/payments later.",
        "brands": {
            "poet": "Популярный поэт (взрослые)",
            "kids": "Идея (детские)",
        },
        "poet": poet,
        "kids": kids,
        "totals": {
            "poet_groups": len(poet["groups"]),
            "poet_persons": len(poet["persons"]),
            "poet_enrollments": len(poet["enrollments"]),
            "kids_groups": len(kids["groups"]),
            "kids_persons": len(kids["persons"]),
            "kids_enrollments": len(kids["enrollments"]),
            "poet_active": sum(1 for e in poet["enrollments"] if e.get("status") == "active"),
            "poet_ended": sum(1 for e in poet["enrollments"] if e.get("status") == "ended"),
            "kids_active": sum(1 for e in kids["enrollments"] if e.get("status") == "active"),
            "kids_ended": sum(1 for e in kids["enrollments"] if e.get("status") == "ended"),
        },
    }
    out_path = ROOT / "scripts" / "data" / "real-tables-phase1.json"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(out["totals"], ensure_ascii=False, indent=2))
    print(f"wrote {out_path}")


if __name__ == "__main__":
    main()
